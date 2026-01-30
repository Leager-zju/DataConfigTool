import os
from pathlib import Path
import openpyxl
import yaml

import pytest

from utils.models import ColumnDef, ConfigTable
from utils.enums import KeyType
from utils.storage import save_table, load_table
from utils.sync import sync_excel_to_yaml


def make_excel(path: Path, sheet_name: str, columns, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # header
    for i, col in enumerate(columns, start=1):
        ws.cell(row=1, column=i, value=col)
    # types (dummy)
    for i, _ in enumerate(columns, start=1):
        ws.cell(row=2, column=i, value="int")
    # key info
    ws.cell(row=3, column=1, value=f"KEY({KeyType.GROUP.value.upper()})")

    for r_idx, row in enumerate(rows, start=4):
        for c_idx, col in enumerate(columns, start=1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(col))

    wb.save(path)


def test_sync_diff_add_delete(tmp_path):
    # prepare YAML table with ids 1,2,3
    table = ConfigTable(table_name="t", group_name="g", key_type=KeyType.GLOBAL)
    table.columns = [ColumnDef(name="id", type="int"), ColumnDef(name="val", type="int")]
    table.data = [{"id": 1, "val": 10}, {"id": 2, "val": 20}, {"id": 3, "val": 30}]

    table_path = tmp_path / "t.yaml"
    save_table(table, table_path)

    # create excel: remove id=2, add id=4
    excel_path = tmp_path / "t.xlsx"
    columns = [c.name for c in table.columns]
    rows = [ {"id":1, "val":10}, {"id":3, "val":30}, {"id":4, "val":40} ]
    make_excel(excel_path, "t", columns, rows)

    # run sync
    sync_excel_to_yaml(excel_path, table_path)

    # reload and assert
    t2 = load_table(table_path)
    ids = [r["id"] for r in t2.data]
    assert set(ids) == {1,3,4}
    assert ids == [1,3,4]


def test_sync_duplicate_pk_raises(tmp_path):
    table = ConfigTable(table_name="t2", group_name="g", key_type=KeyType.GLOBAL)
    table.columns = [ColumnDef(name="id", type="int"), ColumnDef(name="val", type="int")]
    table.data = [{"id": 1, "val": 10}]
    table_path = tmp_path / "t2.yaml"
    save_table(table, table_path)

    excel_path = tmp_path / "t2.xlsx"
    columns = [c.name for c in table.columns]
    # duplicate id 1 twice
    rows = [ {"id":1, "val":10}, {"id":1, "val":11} ]
    make_excel(excel_path, "t2", columns, rows)

    print(table_path, excel_path)

    with pytest.raises(ValueError):
        sync_excel_to_yaml(excel_path, table_path)
