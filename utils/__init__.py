#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
配置管理模块

处理YAML和Excel之间的转换，支持强类型数据验证和多级主键约束。

主要功能：
- 配置表的YAML存储和加载
- Excel临时文件的创建和同步
- 强类型数据验证和转换
- 配置表的批量管理
- 支持三种主键约束：TABLE/GROUP/GLOBAL

模块结构：
- enums: 枚举类型定义
- caches: 主键缓存管理
- models: 数据模型（ColumnDef, ConfigTable）
- yaml_handlers: YAML处理和自定义标签
- excel: Excel文件操作和格式化
- storage: 文件存储操作
- sync: Excel到YAML的数据同步

公开接口：
- KeyType: 主键类型枚举
- ColumnDef: 列定义类
- ConfigTable: 配置表类
- 所有存储和同步函数
"""

from pathlib import Path
from openpyxl import Workbook

# 导入所有公开接口
from .enums import KeyType
from .models import ColumnDef, ConfigTable
from .caches import (
    clear_pk_caches,
    get_global_pk_cache,
    get_group_pk_caches,
)
from .storage import (
    load_table,
    save_table,
    create_table,
    get_group_tables,
    get_all_tables,
    load_all_tables_for_validation,
    get_config_dir,
)
from .excel import format_worktable, get_next_primary_key
from .sync import (
    sync_excel_to_yaml,
    sync_excel_to_all_yaml,
)
from .setting_data import SettingData, PathKey as SettingPathKey

# 获取配置目录和临时目录
_config_dir = SettingData.get_instance().get_path(SettingPathKey.DATA_CONFIG_DIR)
_temp_dir = _config_dir / ".cache"
_temp_dir.mkdir(exist_ok=True)


# ============================================================================
# 临时Excel文件管理
# ============================================================================

def create_temp_excel(table_path: Path) -> Path:
    """为单个配置表创建临时Excel文件
    
    如果临时文件已存在则直接返回路径，避免重复创建。
    
    Args:
        table_path: 配置表文件路径
        
    Returns:
        Path: 临时Excel文件路径
        
    Raises:
        Exception: 当配置表加载或Excel创建失败时
    """
    temp_file = _temp_dir / f"{table_path.stem}.xlsx"
    
    # 如果临时文件不存在，则创建新的
    if not temp_file.exists():
        table = load_table(table_path)

        wb = Workbook()
        ws = wb.active
        ws.title = table.table_name

        format_worktable(ws, table)
        wb.save(temp_file)

    return temp_file


def create_temp_excel_for_group(group_name: str) -> Path:
    """为整个组创建包含多个Table的临时Excel文件
    
    Args:
        group_name: 组名
        
    Returns:
        Path: 临时Excel文件路径
        
    Raises:
        ValueError: 当表格不存在或没有配置表时
        Exception: 当Excel创建失败时
    """
    temp_file = _temp_dir / f"{group_name}.xlsx"
    
    # 如果临时文件不存在，则创建新的
    if not temp_file.exists():
        table_files = get_group_tables(group_name)

        if not table_files:
            raise ValueError(f"分组 {group_name} 没有找到任何配置表")

        wb = Workbook()
        wb.remove(wb.active)  # 移除默认工作表

        # 为每个配置表创建工作表
        for table_file in table_files:
            table = load_table(table_file)
            ws = wb.create_sheet(title=table.table_name)
            format_worktable(ws, table)

        wb.save(temp_file)

    return temp_file

# ============================================================================
# 公开API列表
# ============================================================================

__all__ = [
    # 类型和枚举
    'KeyType',
    'ColumnDef',
    'ConfigTable',
    
    # 缓存操作
    'clear_pk_caches',
    'get_global_pk_cache',
    'get_group_pk_caches',
    
    # 存储操作
    'load_table',
    'save_table',
    'create_table',
    'get_group_tables',
    'get_all_tables',
    'load_all_tables_for_validation',
    'get_config_dir',
    
    # Excel操作
    'format_worktable',
    'get_next_primary_key',
    'create_temp_excel',
    'create_temp_excel_for_group',
    
    # 数据同步
    'sync_excel_to_yaml',
    'sync_excel_to_all_yaml',
]
