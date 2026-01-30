#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
数据同步模块

负责Excel和YAML之间的数据同步，包括数据验证和自动填充。
"""

from typing import List
import openpyxl
from pathlib import Path

from .storage import load_table, save_table

def sync_excel_to_yaml_internal(ws, table_path: Path):
    """将Excel工作表的修改同步到YAML文件（内部函数）
    
    从Excel工作表读取数据，验证后更新对应的YAML文件。
    跳过完全空白的行。
    如果第一列（主键）为空，自动填充为上一行+1。
    
    Excel格式：
    - 第1行：列名
    - 第2行：数据类型
    - 第3行：主键信息（KEY(type)）
    - 第4行以后：数据行
    
    Args:
        ws: openpyxl工作表对象
        table_path: 目标YAML文件路径
        
    Raises:
        Exception: 当数据验证、主键冲突或文件保存失败时
    """
    # 加载原始配置表结构
    table = load_table(table_path)
 
    if not table.columns:
        raise ValueError(f"配置表 {table.table_name} 没有列定义")

    pk_column_name = table.columns[0].name  # 第一列总是主键

    # 读取Excel中的列名（第一行）
    column_names = []
    for col_idx in range(1, ws.max_column + 1):
        col_name = ws.cell(row=1, column=col_idx).value
        if col_name:
            column_names.append(col_name)

    # 读取数据行并验证
    new_pk_list : List[int] = []
    new_datas : List[dict] = []  # 存放 validated_rows
    for row_idx in range(4, ws.max_row + 1):
        row_data = {}
        is_empty_row = True

        # 读取每列的值
        for col_idx, col_name in enumerate(column_names, start=1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None and value != "":
                is_empty_row = False
            row_data[col_name] = table.columns[col_idx - 1].validate_value(value)

        # 跳过空行
        if is_empty_row:
            continue

        # 处理主键自动填充（第一列）
        pk_value = row_data.get(pk_column_name)
        if pk_value is None:
            if new_datas:
                last_pk = new_datas[-1].get(pk_column_name, 0)
                row_data[pk_column_name] = last_pk + 1
            else:
                row_data[pk_column_name] = 1

            # 更新Excel单元格
            try:
                ws.cell(row=row_idx, column=1, value=row_data[pk_column_name])
            except ValueError:
                pass

        new_pk_list.append(row_data[pk_column_name])
        new_datas.append(row_data)

    from collections import Counter
    counts = Counter(new_pk_list)
    for pk_val, cnt in counts.items():
        if pk_val is None or pk_val == "":
            continue
        if cnt > 1:
            raise ValueError(f"主键冲突：主键值 {pk_val} 在Excel中出现了 {cnt} 次")

    old_pk_set = set()
    for r in table.data:
        v = r.get(pk_column_name)

        if v is not None and v != "":
            old_pk_set.add(v)

    new_pk_set = {pk for pk in new_pk_list if pk is not None and pk != ""}

    deleted_pks = old_pk_set - new_pk_set
    added_pks = new_pk_set - old_pk_set

    # 调整缓存：从缓存中移除已删除的主键（以便后续新增验证不会被旧值影响）
    from .pk_cache import apply_pk_diff
    apply_pk_diff(table.group_name, table.table_name, deleted_pks, added_pks, table.key_type)

    # 一切验证通过，使用Excel中的顺序作为最终数据并保存。
    table.data = new_datas
    save_table(table, table_path)


def sync_excel_to_yaml(excel_path: Path, table_path: Path):
    """将单个配置表的Excel修改同步到YAML
    
    Args:
        excel_path: Excel文件路径
        table_path: 目标YAML文件路径
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    sync_excel_to_yaml_internal(ws, table_path)


def sync_excel_to_all_yaml(excel_path: Path, group_name: str):
    """将多个Excel文件的修改同步到整个分组的所有配置表文件
    
    Args:
        excel_path: Excel文件路径
        group_name: 分组名称
    """
    from .storage import get_config_dir
    
    wb = openpyxl.load_workbook(excel_path)
    group_path = get_config_dir() / group_name
    
    for ws in wb.sheetnames:
        table_path = group_path / f"{ws}.yaml"
        if not table_path.exists():
            print(f"警告：文件 {table_path} 不存在，跳过同步")
            continue

        sync_excel_to_yaml_internal(wb[ws], table_path)
