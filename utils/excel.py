#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel处理模块

负责Excel文件的格式化、样式设置和主键处理。
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import ConfigTable


# 定义专业样式
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TYPE_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
PK_HEADER_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 浅红色表示主键列
PK_TYPE_FILL = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
KEY_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # 浅黄色表示主键信息
PK_KEY_FILL = PatternFill(start_color="FFEB99", end_color="FFEB99", fill_type="solid")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


def format_worktable(ws, table: ConfigTable):
    """格式化Excel工作表，设置样式和数据
    
    为工作表应用专业的样式，包括：
    - 表头样式（蓝色背景，白色字体）
    - 类型行样式（浅蓝色背景）
    - 主键信息行（显示主键约束类型）
    - 第一列标记为主键列（浅红色背景）
    - 数据格式化（复杂类型转JSON）
    - 自动调整列宽
    
    Excel格式：
    - 第1行：列名
    - 第2行：数据类型
    - 第3行：主键信息（如 KEY(GLOBAL), KEY(GROUP), KEY(TABLE)）
    - 第4行以后：数据行
    
    Args:
        ws: openpyxl工作表对象
        table: 配置表数据
    """
    # 写入表头（第一行：列名，第二行：数据类型，第三行：主键信息）
    for col_idx, col in enumerate(table.columns, start=1):
        # 第一行：列名
        cell = ws.cell(row=1, column=col_idx, value=col.name)
        cell.font = HEADER_FONT
        cell.fill = PK_HEADER_FILL if col_idx == 1 else HEADER_FILL
        cell.alignment = CENTER_ALIGN

        # 第二行：数据类型
        cell = ws.cell(row=2, column=col_idx, value=col.type)
        cell.fill = PK_TYPE_FILL if col_idx == 1 else TYPE_FILL
        cell.alignment = CENTER_ALIGN
        
        # 第三行：主键信息（只在第一列显示）
        if col_idx == 1:
            key_info = f"KEY({table.key_type.value.upper()})"
            cell = ws.cell(row=3, column=col_idx, value=key_info)
            cell.fill = PK_KEY_FILL
            cell.alignment = CENTER_ALIGN
        else:
            cell = ws.cell(row=3, column=col_idx, value="")
            cell.fill = KEY_FILL

    # 写入数据行（从第四行开始）
    for row_idx, row_data in enumerate(table.data, start=4):
        for col_idx, col in enumerate(table.columns, start=1):
            value = row_data.get(col.name, "")

            # 复杂类型转换为JSON字符串以便在Excel中显示
            if (col.type.startswith("List") or col.type.startswith("Dictionary")) and value is not None and value != "":
                value = json.dumps(value, ensure_ascii=False)
                
            ws.cell(row=row_idx, column=col_idx, value=value)

    # 自动调整列宽以提高可读性
    for col_idx in range(1, len(table.columns) + 1):
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 15


def get_next_primary_key(table: ConfigTable) -> int:
    """获取下一个主键值
    
    主键总是第一列，计算下一个主键值为：最大主键值 + 1
    
    Args:
        table: 配置表对象
        
    Returns:
        int: 下一个主键值
    """
    if not table.columns or table.columns[0].type != "int":
        return None
    
    if not table.data:
        return 1
    
    # 找到最大的主键值（第一列）
    pk_column_name = table.columns[0].name
    max_pk = 0
    for row in table.data:
        pk_value = row.get(pk_column_name, 0)
        if isinstance(pk_value, int) and pk_value > max_pk:
            max_pk = pk_value
    
    return max_pk + 1
