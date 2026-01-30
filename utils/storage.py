#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
存储操作模块

负责配置表文件的加载、保存和查询。
"""

import yaml
from pathlib import Path
from typing import Dict, List
from pathlib import Path
from openpyxl import Workbook

from .excel import format_worktable
from .models import ConfigTable, ColumnDef
from .enums import KeyType
from .pk_cache import clear_pk_caches
from .setting_data import SettingData, PathKey as SettingPathKey


# 获取配置目录和临时文件目录
_config_dir = SettingData.get_instance().get_path(SettingPathKey.DATA_CONFIG_DIR)
_temp_dir = _config_dir / ".cache"
_temp_dir.mkdir(exist_ok=True)


def get_config_dir() -> Path:
    """获取配置目录路径
    
    Returns:
        Path: 配置目录路径
    """
    return _config_dir


def load_table(table_path: Path) -> ConfigTable:
    """从YAML文件加载单个配置表
    
    加载完成后会自动更新主键缓存。
    
    Args:
        table_path: YAML文件路径
        
    Returns:
        ConfigTable: 加载的配置表对象
        
    Raises:
        FileNotFoundError: 当文件不存在时
        yaml.YAMLError: 当YAML格式错误时
        ValueError: 当主键验证失败时
    """
    with open(table_path, "r", encoding="utf-8") as f:
        data = yaml.load(f, Loader=yaml.SafeLoader)

    table = ConfigTable.from_dict(data)
    
    return table


def save_table(table: ConfigTable, table_path: Path):
    """保存配置表到YAML文件
    
    保存后会自动更新主键缓存。
    
    Args:
        table: 要保存的配置表对象
        table_path: 目标YAML文件路径
        
    Raises:
        IOError: 当文件写入失败时
    """
    # 确保目录存在
    table_path.parent.mkdir(parents=True, exist_ok=True)
    
    with open(table_path, "w", encoding="utf-8") as f:
        yaml.dump(table.to_dict(), f, allow_unicode=True, sort_keys=False)


def create_table(table_path: Path, table_name: str, group_name: str, columns: List[ColumnDef], key_type: KeyType = KeyType.GROUP):
    """创建新的配置表
    
    Args:
        table_path: 新配置表的文件路径
        table_name: 配置表名称
        group_name: 所属组名称
        columns: 列定义列表
        key_type: 主键类型
    """
    table = ConfigTable(table_name=table_name, group_name=group_name, columns=columns, key_type=key_type)
    save_table(table, table_path)


def get_group_tables(group_name: str) -> List[Path]:
    """获取指定组的所有配置表文件
    
    Args:
        group_name: 组名
        
    Returns:
        List[Path]: 排序后的配置表文件路径列表
    """
    table_path = _config_dir / group_name
    if not table_path.exists():
        return []
    
    pattern = "*.yaml"
    return sorted(table_path.glob(pattern))


def get_all_tables() -> Dict[str, List[Path]]:
    """获取所有组别及其配置表文件

    扫描配置目录，按组返回所有配置表文件。
    
    Returns:
        Dict[str, List[Path]]: 组名到配置表文件列表的映射
    """
    tables = {}
    
    # 搜索所有YAML文件
    yaml_files = list(_config_dir.rglob("*.yaml")) + list(_config_dir.rglob("*.yml"))

    for yaml_file in yaml_files:
        # 使用父目录名作为组名
        group_name = yaml_file.parent.name
        
        # 跳过缓存目录
        if group_name == ".cache":
            continue

        if group_name not in tables:
            tables[group_name] = []
        tables[group_name].append(yaml_file)

    return tables


def load_all_tables_for_validation():
    """加载所有表以进行跨表/跨分组的主键验证
    
    用于初始化全局和分组主键缓存，并检查所有表的主键约束。
    
    Returns:
        List[str]: 主键冲突的错误信息列表（空列表表示无冲突）
    """
    clear_pk_caches()
    tables = get_all_tables()
    conflicts = []
    
    for group_name in tables.keys():
        table_files = tables[group_name]
        for table_file in table_files:
            try:
                with open(table_file, "r", encoding="utf-8") as f:
                    data = yaml.load(f, Loader=yaml.SafeLoader)
                table = ConfigTable.from_dict(data)
                
                # 根据表的key_type验证主键约束
                table.validate_all_primary_keys()
            except Exception as e:
                error_msg = f"验证表 {table_file} 主键时出错：{e}"
                print(f"错误：{error_msg}")
                conflicts.append(error_msg)
    
    return conflicts

# ============================================================================
# 临时Excel文件管理
# ============================================================================

def create_temp_excel(table_path: Path) -> Path:
    """为单个配置表创建临时Excel文件
    
    如果临时文件已存在则直接返回路径，避免重复创建。
    
    Args:
        table_path: 配置表文件路径
        
    Returns:
        Path: 临时Excel文件路径
        
    Raises:
        Exception: 当配置表加载或Excel创建失败时
    """
    temp_file = _temp_dir / f"{table_path.stem}.xlsx"
    
    # 如果临时文件不存在，则创建新的
    if not temp_file.exists():
        table = load_table(table_path)

        wb = Workbook()
        ws = wb.active
        ws.title = table.table_name

        format_worktable(ws, table)
        wb.save(temp_file)

    return temp_file


def create_temp_excel_for_group(group_name: str) -> Path:
    """为整个组创建包含多个Table的临时Excel文件
    
    Args:
        group_name: 组名
        
    Returns:
        Path: 临时Excel文件路径
        
    Raises:
        ValueError: 当表格不存在或没有配置表时
        Exception: 当Excel创建失败时
    """
    temp_file = _temp_dir / f"{group_name}.xlsx"
    
    # 如果临时文件不存在，则创建新的
    if not temp_file.exists():
        table_files = get_group_tables(group_name)

        if not table_files:
            raise ValueError(f"分组 {group_name} 没有找到任何配置表")

        wb = Workbook()
        wb.remove(wb.active)  # 移除默认工作表

        # 为每个配置表创建工作表
        for table_file in table_files:
            table = load_table(table_file)
            ws = wb.create_sheet(title=table.table_name)
            format_worktable(ws, table)

        wb.save(temp_file)

    return temp_file